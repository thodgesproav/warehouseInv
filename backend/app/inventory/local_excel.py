from __future__ import annotations

import fcntl
import hashlib
import os
import shutil
import tempfile
import threading
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import TableColumn

from ..config import settings
from ..database import get_mapping
from .base import InsufficientStock, InventoryProvider, ItemNotFound, StockConflict, SyncUnavailable


REQUIRED_COLUMNS = ("Inventory ID", "On Order", "Quantity On Order", "Image")


class LocalExcelInventoryProvider(InventoryProvider):
    def __init__(self, path: Path | None = None, sheet_name: str | None = None):
        self.path = Path(path or settings.workbook_path)
        self.sheet_name = sheet_name or settings.inventory_sheet
        self._thread_lock = threading.RLock()
        self._last_sync: str | None = None
        self._last_error: str | None = None

    @contextmanager
    def _locked(self) -> Iterator[None]:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        lock_path = self.path.with_suffix(self.path.suffix + ".lock")
        with self._thread_lock, open(lock_path, "a+b") as lock:
            fcntl.flock(lock.fileno(), fcntl.LOCK_EX)
            try:
                yield
            finally:
                fcntl.flock(lock.fileno(), fcntl.LOCK_UN)

    def _open(self):
        if not self.path.exists():
            raise SyncUnavailable(f"Workbook not found: {self.path}")
        wb = load_workbook(self.path)
        if self.sheet_name not in wb.sheetnames:
            wb.close()
            raise SyncUnavailable(f"Worksheet not found: {self.sheet_name}")
        return wb, wb[self.sheet_name]

    @staticmethod
    def _headers(ws) -> tuple[int, dict[str, int]]:
        for row in range(1, min(ws.max_row, 25) + 1):
            values = [str(ws.cell(row, col).value).strip() if ws.cell(row, col).value is not None else "" for col in range(1, ws.max_column + 1)]
            if "SOH" in values and ("Description" in values or "Master/Part No." in values):
                return row, {value: index + 1 for index, value in enumerate(values) if value}
        raise SyncUnavailable("Could not identify the inventory header row")

    def _backup(self) -> None:
        settings.backup_path.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
        shutil.copy2(self.path, settings.backup_path / f"{self.path.stem}-{stamp}{self.path.suffix}")

    def _save(self, wb) -> None:
        self._backup()
        fd, temporary = tempfile.mkstemp(prefix="inventory-", suffix=".xlsx", dir=self.path.parent)
        os.close(fd)
        try:
            wb.save(temporary)
            os.replace(temporary, self.path)
        finally:
            if os.path.exists(temporary):
                os.unlink(temporary)

    def prepare_workbook(self) -> None:
        with self._locked():
            wb, ws = self._open()
            try:
                header_row, headers = self._headers(ws)
                changed = False
                for name in REQUIRED_COLUMNS:
                    if name not in headers:
                        column = ws.max_column + 1
                        ws.cell(header_row, column, name)
                        source = ws.cell(header_row, column - 1)
                        ws.cell(header_row, column)._style = source._style
                        headers[name] = column
                        changed = True
                id_col = headers["Inventory ID"]
                seen: set[str] = set()
                for row in range(header_row + 1, ws.max_row + 1):
                    if not any(ws.cell(row, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
                        continue
                    value = str(ws.cell(row, id_col).value or "").strip()
                    if not value or value in seen:
                        seed = "|".join(str(ws.cell(row, headers.get(k, 1)).value or "") for k in ("Master/Part No.", "Description", "Bin Number"))
                        base = "INV-" + hashlib.sha1(seed.encode()).hexdigest()[:8].upper()
                        value = base
                        counter = 2
                        while value in seen:
                            value = f"{base}-{counter}"
                            counter += 1
                        ws.cell(row, id_col, value)
                        changed = True
                    seen.add(value)
                if ws.tables:
                    table = next(iter(ws.tables.values()))
                    table.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
                    table.tableColumns = [TableColumn(id=index, name=ws.cell(header_row, index).value) for index in range(1, ws.max_column + 1)]
                if changed:
                    self._save(wb)
            finally:
                wb.close()

    @staticmethod
    def _truthy(value: Any) -> bool:
        return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on order", "ordered"}

    def _product(self, ws, row: int, headers: dict[str, int], mapping: dict[str, str]) -> dict[str, Any]:
        raw = {name: ws.cell(row, column).value for name, column in headers.items()}
        def value(key: str, default: Any = "") -> Any:
            return raw.get(mapping.get(key, ""), default)
        stock_value = value("stock", 0)
        try: stock = int(stock_value or 0)
        except (ValueError, TypeError): stock = 0
        quantity_ordered = value("quantity_on_order", 0)
        try: quantity_ordered = int(quantity_ordered or 0)
        except (ValueError, TypeError): quantity_ordered = 0
        return {
            "id": str(value("id") or "").strip(), "name": str(value("name") or "Unnamed item").strip(),
            "manufacturer": str(value("manufacturer") or "").strip(), "model": str(value("model") or "").strip(),
            "sku": str(value("sku") or "").strip(), "stock": stock, "location": str(value("location") or "").strip(),
            "on_order": self._truthy(value("on_order")), "quantity_on_order": quantity_ordered,
            "image": str(value("image") or "").strip(), "category": str(value("category") or "").strip(),
            "discontinued": str(value('discontinued', '')).strip().lower() in {'1','true','yes','y','x','discontinued'},
            "raw_fields": raw,
        }

    def get_inventory(self, force: bool = False) -> list[dict[str, Any]]:
        del force
        with self._locked():
            wb, ws = self._open()
            try:
                header_row, headers = self._headers(ws)
                mapping = get_mapping()
                products = [self._product(ws, row, headers, mapping) for row in range(header_row + 1, ws.max_row + 1) if any(ws.cell(row, c).value not in (None, "") for c in range(1, ws.max_column + 1))]
                self._last_sync = datetime.now(timezone.utc).isoformat()
                self._last_error = None
                return products
            except Exception as exc:
                self._last_error = str(exc)
                raise
            finally:
                wb.close()

    def _find(self, ws, headers: dict[str, int], header_row: int, item_id: str) -> int:
        id_header = get_mapping().get("id", "Inventory ID")
        if id_header not in headers: raise ItemNotFound(item_id)
        for row in range(header_row + 1, ws.max_row + 1):
            if str(ws.cell(row, headers[id_header]).value or "").strip() == item_id: return row
        raise ItemNotFound(item_id)

    def adjust_stock(self, item_id: str, quantity: int, expected_current_soh: int) -> dict[str, Any]:
        if quantity == 0: raise ValueError("Quantity cannot be zero")
        with self._locked():
            wb, ws = self._open()
            try:
                header_row, headers = self._headers(ws); mapping = get_mapping(); row = self._find(ws, headers, header_row, item_id)
                stock_header = mapping.get("stock", "SOH")
                if stock_header not in headers: raise SyncUnavailable("The mapped stock column is missing")
                old = int(ws.cell(row, headers[stock_header]).value or 0)
                if old != expected_current_soh: raise StockConflict(f"Stock changed from {expected_current_soh} to {old}")
                new = old + quantity
                if new < 0: raise InsufficientStock(f"Only {old} are currently available")
                ws.cell(row, headers[stock_header], new)
                self._save(wb)
                self._last_sync = datetime.now(timezone.utc).isoformat(); self._last_error = None
                product = self._product(ws, row, headers, mapping); product["old_stock"] = old
                return product
            finally: wb.close()

    def update_item(self, item_id: str, fields: dict[str, Any]) -> dict[str, Any]:
        with self._locked():
            wb, ws = self._open()
            try:
                header_row, headers = self._headers(ws); row = self._find(ws, headers, header_row, item_id)
                for header, value in fields.items():
                    if header in headers and header != get_mapping().get("id"): ws.cell(row, headers[header], value)
                self._save(wb)
                return self._product(ws, row, headers, get_mapping())
            finally: wb.close()

    def add_item(self, fields: dict[str, Any]) -> dict[str, Any]:
        with self._locked():
            wb, ws = self._open()
            try:
                header_row, headers = self._headers(ws); row = ws.max_row + 1
                item_id = str(fields.get("Inventory ID") or f"INV-{hashlib.sha1(os.urandom(16)).hexdigest()[:8].upper()}")
                fields["Inventory ID"] = item_id
                for header, value in fields.items():
                    if header in headers: ws.cell(row, headers[header], value)
                if ws.tables:
                    table = next(iter(ws.tables.values())); table.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{row}"
                self._save(wb)
                return self._product(ws, row, headers, get_mapping())
            finally: wb.close()

    def delete_item(self, item_id: str) -> None:
        with self._locked():
            wb, ws = self._open()
            try:
                header_row, headers = self._headers(ws); row = self._find(ws, headers, header_row, item_id)
                ws.delete_rows(row, 1)
                if ws.tables:
                    table = next(iter(ws.tables.values())); table.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
                self._save(wb)
            finally: wb.close()

    def get_columns(self) -> list[str]:
        with self._locked():
            wb, ws = self._open()
            try: return list(self._headers(ws)[1])
            finally: wb.close()

    def get_sync_status(self) -> dict[str, Any]:
        return {"provider": "local_excel", "ok": self._last_error is None, "last_sync": self._last_sync, "error": self._last_error, "workbook": str(self.path)}
