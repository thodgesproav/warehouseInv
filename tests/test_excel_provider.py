from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor

import pytest
from openpyxl import load_workbook

from app.inventory.base import InsufficientStock, StockConflict


def test_reads_supplied_workbook_and_mapping(provider):
    items = provider.get_inventory()
    assert len(items) == 98
    assert items[0]["name"].startswith("50cm Super Slim")
    assert items[0]["manufacturer"] == "Comsol"
    assert items[0]["stock"] == 0
    assert all(item["id"].startswith("INV-") for item in items)


def test_required_columns_are_added_without_losing_unknown_columns(provider):
    columns = provider.get_columns()
    assert {"Inventory ID", "On Order", "Quantity On Order", "Image"}.issubset(columns)
    assert {"Notes", "ALT/OLD SKU STOCK", "OLD SOH"}.issubset(columns)


def test_column_reorder_and_new_unknown_column_are_tolerated(provider):
    wb = load_workbook(provider.path); ws = wb["Warehouse"]
    ws.insert_cols(1); ws.cell(1, 1, "Project Code")
    for row in range(2, ws.max_row + 1): ws.cell(row, 1, f"P-{row}")
    wb.save(provider.path); wb.close()
    items = provider.get_inventory()
    assert len(items) == 98
    assert items[0]["raw_fields"]["Project Code"] == "P-2"
    provider.update_item(items[0]["id"], {"Description": "Updated description"})
    assert provider.get_inventory()[0]["raw_fields"]["Project Code"] == "P-2"


def test_row_reorder_keeps_identity(provider):
    before = provider.get_inventory(); target = before[3]
    wb = load_workbook(provider.path); ws = wb["Warehouse"]
    values = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column)]
    values.reverse()
    for r, values_row in enumerate(values, 2):
        for c, value in enumerate(values_row, 1): ws.cell(r, c, value)
    wb.save(provider.path); wb.close()
    found = next(item for item in provider.get_inventory() if item["id"] == target["id"])
    assert found["name"] == target["name"]


def test_stock_deduction_and_negative_prevention(provider):
    item = next(x for x in provider.get_inventory() if x["stock"] >= 2)
    updated = provider.adjust_stock(item["id"], -2, item["stock"])
    assert updated["stock"] == item["stock"] - 2
    with pytest.raises(InsufficientStock): provider.adjust_stock(item["id"], -(updated["stock"] + 1), updated["stock"])


def test_expected_stock_prevents_conflict(provider):
    item = next(x for x in provider.get_inventory() if x["stock"] >= 1)
    provider.adjust_stock(item["id"], -1, item["stock"])
    with pytest.raises(StockConflict): provider.adjust_stock(item["id"], -1, item["stock"])


def test_simultaneous_updates_only_one_wins(provider):
    item = next(x for x in provider.get_inventory() if x["stock"] >= 2)
    def take():
        try: return provider.adjust_stock(item["id"], -2, item["stock"])["stock"]
        except StockConflict: return "conflict"
    with ThreadPoolExecutor(max_workers=2) as pool: outcomes = list(pool.map(lambda _: take(), range(2)))
    assert outcomes.count("conflict") == 1
    assert next(x for x in provider.get_inventory() if x["id"] == item["id"])["stock"] == item["stock"] - 2


def test_add_edit_delete_and_manual_reload(provider):
    created = provider.add_item({"Description": "Test adapter", "SOH": 4, "Project Code": "ignored"})
    assert created["stock"] == 4
    provider.update_item(created["id"], {"SOH": 7, "Notes": "Manual field preserved"})
    assert next(x for x in provider.get_inventory() if x["id"] == created["id"])["stock"] == 7
    provider.delete_item(created["id"])
    assert all(x["id"] != created["id"] for x in provider.get_inventory())


def test_removing_optional_column_does_not_break_read(provider):
    wb = load_workbook(provider.path); ws = wb["Warehouse"]
    headers = [cell.value for cell in ws[1]]; ws.delete_cols(headers.index("OLD SOH") + 1)
    wb.save(provider.path); wb.close()
    assert len(provider.get_inventory()) == 98

